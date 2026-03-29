import json
from io import BytesIO
from datetime import datetime
from typing import List

from fastapi import BackgroundTasks, Body, Depends, FastAPI, File, HTTPException, Path, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from fastapi.security import OAuth2PasswordRequestForm
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from . import auth, crud, email_service, models, schemas
from .database import engine, get_db
from .schema_init import ensure_schema

models.Base.metadata.create_all(bind=engine)
ensure_schema(engine)

app = FastAPI(title="Finance API")


def _handle_value_error(exc: ValueError):
    raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.get("/health")
def health():
    return {"status": "ok"}


@app.post("/register/request")
def request_registration(
    background_tasks: BackgroundTasks,
    email: str = Body(...),
    password: str = Body(...),
    name: str = Body(...),
    db: Session = Depends(get_db),
):
    if crud.get_user_by_email(db, email):
        raise HTTPException(status_code=400, detail="Email уже зарегистрирован")

    code = crud.create_email_verification(db, email, password, name)
    background_tasks.add_task(email_service.send_code, email, code)
    return {"detail": "Код отправлен на email"}


@app.post("/register/confirm")
def confirm_registration(
    email: str = Body(...),
    code: str = Body(...),
    db: Session = Depends(get_db),
):
    from .models import EmailVerification

    record = db.query(EmailVerification).filter(
        EmailVerification.email == email,
        EmailVerification.code == code,
        EmailVerification.expires_at > datetime.utcnow(),
    ).first()

    if not record:
        raise HTTPException(status_code=400, detail="Неверный код или срок действия истёк")

    user_data = schemas.UserCreate(
        email=record.email,
        password=record.password,
        name=record.name,
        is_verified=True,
    )
    user = crud.create_user(db, user_data)

    db.delete(record)
    db.commit()

    return {"detail": "Регистрация успешна", "user_id": user.id}


@app.get("/profile", response_model=schemas.UserOut)
def get_profile(current_user=Depends(auth.get_current_user)):
    return current_user


@app.post("/token", response_model=schemas.Token)
def login_for_access_token(
    form_data: OAuth2PasswordRequestForm = Depends(),
    db: Session = Depends(get_db),
):
    user = auth.authenticate_user(db, form_data.username, form_data.password)
    if not user:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Incorrect credentials")
    access_token = auth.create_access_token(data={"sub": user.email})
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/statistics", response_model=schemas.StatisticsOut)
def get_statistics(
    period: str = Query("month"),
    start_date: str = Query(None),
    end_date: str = Query(None),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    start = None
    end = None

    if start_date:
        try:
            start = datetime.fromisoformat(start_date).date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Неверный формат start_date (используйте YYYY-MM-DD)") from exc

    if end_date:
        try:
            end = datetime.fromisoformat(end_date).date()
        except ValueError as exc:
            raise HTTPException(status_code=400, detail="Неверный формат end_date (используйте YYYY-MM-DD)") from exc

    return crud.get_user_statistics(
        db=db,
        owner_id=current_user.id,
        period=period,
        start_date=start,
        end_date=end,
    )


@app.post("/categories", response_model=schemas.CategoryOut)
def create_category(
    category: schemas.CategoryCreate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        return crud.create_category(db, user_id=current_user.id, category=category)
    except ValueError as exc:
        _handle_value_error(exc)


@app.get("/categories", response_model=List[schemas.CategoryOut])
def list_categories(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return crud.get_categories_for_user(db, user_id=current_user.id)


@app.post("/accounts", response_model=schemas.AccountOut)
def create_account(
    account: schemas.AccountCreate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return crud.create_account(db, user_id=current_user.id, account=account)


@app.get("/accounts", response_model=List[schemas.AccountOut])
def list_accounts(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return crud.get_accounts_for_user(db, user_id=current_user.id)


@app.put("/accounts/{account_id}", response_model=schemas.AccountOut)
def update_account(
    account_id: int,
    account: schemas.AccountUpdate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    updated = crud.update_account(db, user_id=current_user.id, account_id=account_id, account=account)
    if not updated:
        raise HTTPException(status_code=404, detail="Account not found")
    return updated


@app.post("/budgets", response_model=schemas.BudgetOut)
def create_budget(
    budget: schemas.BudgetCreate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        return crud.create_budget(db, user_id=current_user.id, budget=budget)
    except ValueError as exc:
        _handle_value_error(exc)


@app.get("/budgets", response_model=List[schemas.BudgetOut])
def list_budgets(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return crud.get_budgets_for_user(db, user_id=current_user.id)


@app.put("/budgets/{budget_id}", response_model=schemas.BudgetOut)
def update_budget(
    budget_id: int,
    budget: schemas.BudgetUpdate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        updated = crud.update_budget(db, user_id=current_user.id, budget_id=budget_id, budget=budget)
    except ValueError as exc:
        _handle_value_error(exc)
    if not updated:
        raise HTTPException(status_code=404, detail="Budget not found")
    return updated


@app.get("/transactions/export")
def export_transactions(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    rows = crud.export_transactions(db, owner_id=current_user.id)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Transactions"
    sheet.append(
        ["amount", "category_id", "category_name", "account_id", "note", "type", "date"]
    )
    for row in rows:
        sheet.append(
            [
                row["amount"],
                row["category_id"],
                row["category_name"],
                row["account_id"],
                row["note"],
                row["type"],
                row["date"],
            ]
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="transactions.xlsx"'},
    )


@app.post("/transactions/import")
async def import_transactions(
    file: UploadFile = File(...),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        content = await file.read()
        filename = (file.filename or "").lower()
        if filename.endswith(".xlsx"):
            workbook = load_workbook(filename=BytesIO(content), data_only=True)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Empty file")
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            raw_data = []
            for row in rows[1:]:
                if row is None or all(cell in (None, "") for cell in row):
                    continue
                item = {}
                for index, header in enumerate(headers):
                    if not header:
                        continue
                    item[header] = row[index] if index < len(row) else None
                raw_data.append(item)
        else:
            raw_data = json.loads(content)
        data = [schemas.TransactionExport(**item) for item in raw_data]
        crud.import_transactions(db, owner_id=current_user.id, data=data)
        return {"message": "Импорт выполнен"}
    except ValueError as exc:
        _handle_value_error(exc)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Ошибка импорта: {exc}") from exc


@app.post("/transactions", response_model=schemas.TransactionOut)
def create_transaction(
    tx: schemas.TransactionCreate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        return crud.create_transaction(db, owner_id=current_user.id, tx=tx)
    except ValueError as exc:
        _handle_value_error(exc)


@app.get("/transactions", response_model=List[schemas.TransactionOut])
def list_transactions(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return crud.get_transactions_for_user(db, owner_id=current_user.id)


@app.put("/transactions/{tx_id}", response_model=schemas.TransactionOut)
def update_transaction(
    tx_id: int,
    tx: schemas.TransactionUpdate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        updated = crud.update_transaction(db, owner_id=current_user.id, tx_id=tx_id, tx=tx)
    except ValueError as exc:
        _handle_value_error(exc)
    if not updated:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return updated


@app.delete("/transactions/{tx_id}", status_code=204)
def delete_transaction(
    tx_id: int = Path(..., description="ID транзакции"),
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    deleted = crud.delete_transaction(db, owner_id=current_user.id, tx_id=tx_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return


@app.post("/debts", response_model=schemas.DebtOut)
def create_debt(
    debt: schemas.DebtCreate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        return crud.create_debt(db, user_id=current_user.id, debt=debt)
    except ValueError as exc:
        _handle_value_error(exc)


@app.get("/debts", response_model=List[schemas.DebtOut])
def list_debts(
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    return crud.get_debts_for_user(db, user_id=current_user.id)


@app.put("/debts/{debt_id}", response_model=schemas.DebtOut)
def update_debt(
    debt_id: int,
    debt: schemas.DebtUpdate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        updated = crud.update_debt(db, user_id=current_user.id, debt_id=debt_id, debt=debt)
    except ValueError as exc:
        _handle_value_error(exc)
    if not updated:
        raise HTTPException(status_code=404, detail="Debt not found")
    return updated


@app.delete("/debts/{debt_id}", status_code=204)
def delete_debt(
    debt_id: int,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    deleted = crud.delete_debt(db, user_id=current_user.id, debt_id=debt_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="Debt not found")
    return


@app.post("/debts/{debt_id}/events", response_model=schemas.DebtEventOut)
def create_debt_event(
    debt_id: int,
    event: schemas.DebtEventCreate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        created = crud.create_debt_event(db, user_id=current_user.id, debt_id=debt_id, event=event)
    except ValueError as exc:
        _handle_value_error(exc)
    if not created:
        raise HTTPException(status_code=404, detail="Debt not found")
    return created


@app.put("/debts/{debt_id}/events/{event_id}", response_model=schemas.DebtEventOut)
def update_debt_event(
    debt_id: int,
    event_id: int,
    event: schemas.DebtEventUpdate,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    try:
        updated = crud.update_debt_event(
            db,
            user_id=current_user.id,
            debt_id=debt_id,
            event_id=event_id,
            event=event,
        )
    except ValueError as exc:
        _handle_value_error(exc)
    if not updated:
        raise HTTPException(status_code=404, detail="Debt event not found")
    return updated


@app.delete("/debts/{debt_id}/events/{event_id}", status_code=204)
def delete_debt_event(
    debt_id: int,
    event_id: int,
    current_user=Depends(auth.get_current_user),
    db: Session = Depends(get_db),
):
    deleted = crud.delete_debt_event(
        db,
        user_id=current_user.id,
        debt_id=debt_id,
        event_id=event_id,
    )
    if not deleted:
        raise HTTPException(status_code=404, detail="Debt event not found")
    return
